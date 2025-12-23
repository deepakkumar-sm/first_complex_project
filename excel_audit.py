import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from datetime import datetime

# load excel files
df_UBR = pd.read_excel("Excel_Audit_Project.xlsx", sheet_name="Sheet1")
df_CVS = pd.read_excel("Excel_Audit_Project.xlsx", sheet_name="Sheet2")

print(df_UBR)
print() 
print(df_CVS)


# Initialize audit columns

df_UBR["Check_Currency_UBR"] = ""
df_UBR["Check_Amount_UBR"] = ""
df_UBR["Check_Course_Term_UBR"] = ""
df_UBR["Check_vendor_name_UBR"] = ""
# df_UBR["Check_deliverydate_conflict_UBR"] = ""
df_UBR["MI_Eligibility"] = ""


# merge UBR with CVS on course name

df_merged = df_UBR.merge(
    df_CVS[["Course_Name","Currency","Amount","Course_Term","vendor_name"]],
    on="Course_Name",
    how="left",
    suffixes=("","_CVS")
)

# helper function for comparison

def compare_with_cvs(row, col, col_cvs):
    if pd.isna(row[col_cvs]):
        return "Course Not Found"
    return "Match" if row[col] == row[col_cvs] else "Mismatch"

# vectorized comparison
df_merged["Check_Currency_UBR"] = df_merged.apply(lambda r: compare_with_cvs(r, "Currency", "Currency_CVS"), axis=1)
df_merged["Check_Amount_UBR"] = df_merged.apply(lambda r: compare_with_cvs(r, "Amount", "Amount_CVS"), axis=1)
df_merged["Check_Course_Term_UBR"] = df_merged.apply(lambda r: compare_with_cvs(r, "Course_Term", "Course_Term_CVS"), axis=1)
df_merged["Check_vendor_name_UBR"] = df_merged.apply(lambda r: compare_with_cvs(r, "vendor_name", "vendor_name_CVS"), axis=1)

df_merged.loc[df_merged["POD_Recd"] == "No", "MI_Eligibility"] = "Not Eligible"
df_merged.loc[df_merged["POD_Recd"] == "Yes", "MI_Eligibility"] = "Eligible"


# save to new excel file
df_merged.to_excel("MI_Audit_UBR_audited.xlsx", index=False)

# styling the excel output fle

wb = load_workbook(filename='MI_Audit_UBR_audited.xlsx')
sheet = wb.active

red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

for row in sheet.iter_rows(min_row=2, min_col=12, max_col=16):

    for cell in row:
        if cell.value == "Mismatch":
            cell.fill = red_fill
        elif cell.value == "Course Not Found":
            cell.fill = yellow_fill


wb.save("MI_Audit_UBR_audited.xlsx")


# Summary detailed report

summary_currency = df_merged["Check_Currency_UBR"].value_counts()
summary_amount = df_merged["Check_Amount_UBR"].value_counts()
summary_term = df_merged["Check_Course_Term_UBR"].value_counts()
summary_vendor_check = df_merged["Check_vendor_name_UBR"].value_counts()
summary_MIEligible = df_merged.loc[df_merged["POD_Recd"] == "Yes", "MI_Eligibility"].value_counts()
summary_MInotEligible = df_merged.loc[df_merged["POD_Recd"] == "No", "MI_Eligibility"].value_counts()


print("Audits Completed. Results saved as MI_Audit_UBR_audited.xlsx")
print("\n---- Summary Report ----")
print("Currency Check: ")
print(summary_currency)
print("\nAmount Check: ")
print(summary_amount)
print("\nTerm Check")
print(summary_term)
print("\nVendor Check")
print(summary_vendor_check)
print()
print(summary_MInotEligible)

# ---------------- HTML Summary Report ----------------

# Audit stats
audit_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
total_records = len(df_merged)
matched_records = (df_merged[["Check_Currency_UBR","Check_Amount_UBR","Check_Course_Term_UBR","Check_vendor_name_UBR"]] == "Match").all(axis=1).sum()
mismatched_records = (df_merged[["Check_Currency_UBR","Check_Amount_UBR","Check_Course_Term_UBR","Check_vendor_name_UBR"]] == "Mismatch").any(axis=1).sum()
course_not_found_records = (df_merged[["Check_Currency_UBR","Check_Amount_UBR","Check_Course_Term_UBR","Check_vendor_name_UBR"]] == "Course Not Found").any(axis=1).sum()
not_eligible_records = (df_merged["MI_Eligibility"] == "Not Eligible").sum()

# Filter tables
df_mismatched = df_merged[(df_merged[["Check_Currency_UBR","Check_Amount_UBR","Check_Course_Term_UBR","Check_vendor_name_UBR"]] == "Mismatch").any(axis=1)]
df_course_not_found = df_merged[(df_merged[["Check_Currency_UBR","Check_Amount_UBR","Check_Course_Term_UBR","Check_vendor_name_UBR"]] == "Course Not Found").any(axis=1)]
df_not_eligible = df_merged[df_merged["MI_Eligibility"] == "Not Eligible"]

# Build HTML
style = """
<style>
body { font-family: Arial, sans-serif; margin: 40px; background: #f9fafb; }
h1, h2 { color: #1f2937; }
.summary { background: #ecf0f1; padding: 16px; border-radius: 8px; margin-bottom: 24px; }
.kpi { margin: 8px 0; }
table { width: 100%; border-collapse: collapse; margin-top: 12px; }
th, td { border: 1px solid #e5e7eb; padding: 8px; text-align: center; }
th { background: #111827; color: #ffffff; }
.mismatched table { background: #efced0; }
.course_not_found table { background: #f1daaf; }
.not_eligible table { background: #ebd8d5; }
</style>
"""

html = f"""
<html>
<head><title>Audit Summary Report</title>{style}</head>
<body>
  <h1>Audit Summary Report</h1>
  <div class="summary">
    <div class="kpi"><strong>Date/Time of Audit:</strong> {audit_date}</div>
    <div class="kpi"><strong>Total Records Audited:</strong> {total_records}</div>
    <div class="kpi"><strong>Mismatched Records:</strong> {mismatched_records}</div>
    <div class="kpi"><strong>Course Not Found:</strong> {course_not_found_records}</div>
    <div class="kpi"><strong>Not Eligible Records:</strong> {not_eligible_records}</div>
  </div>

  <h2>Mismatched Records</h2><h4> - differences found in course amount/currency/term/vendor</h4>
  <div class="mismatched">
    {df_mismatched.to_html(index=False)}
  </div>

  <h2>Course Not Found Records</h2><h4> - course is not found in the base report</h4>
  <div class="course_not_found">
    {df_course_not_found.to_html(index=False)}
  </div>

  <h2>Not Eligible Records</h2><h4> - POD is missing and Invoice hasn't been recieved yet</h4>
  <div class="not_eligible">
    {df_not_eligible.to_html(index=False)}
  </div>
</body>
</html>
"""

# Save HTML file
with open("Audit_Summary_Report.html", "w", encoding="utf-8") as f:
    f.write(html)

print("Audit summary HTML report created: Audit_Summary_Report.html")