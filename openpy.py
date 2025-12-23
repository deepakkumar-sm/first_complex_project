from openpyxl import load_workbook
# from openpyxl import workbook
# import pandas as pd


wb = load_workbook(filename='Excel_Audit_Project.xlsx')
sheet1 = wb['Sheet1']
sheet2 = wb['Sheet2']

print(sheet1.cell(row=2, column=8).value)
print(sheet2.cell(row=2, column=8).value)